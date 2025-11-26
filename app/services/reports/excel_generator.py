from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins


class ExcelReportGenerator:
    """
    Genera reportes Excel para:
        - Un solo host (detalle completo + NVD + gráfica PIE)
        - Múltiples hosts (reporte global en una única hoja)
    """

    COLORS = {
        "header_blue": "003A75FF",
        "header_gray": "00999999",
        "header_red": "00CC3333",
        "header_orange": "00FF8800",
    }

    SEVERITY_COLORS = {
        "CRITICAL": "00FF0000",
        "HIGH": "00FF8800",
        "MEDIUM": "00FFCC00",
        "LOW": "0099CC00",
    }

    # Estilo de encabezado
    def _style_header(self, worksheet, color):
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-ajuste de columnas
    def _auto_fit_columns(self, ws):
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_length = 0
            for cell in col_cells:
                try:
                    val_length = len(str(cell.value)) if cell.value else 0
                    if val_length > max_length:
                        max_length = val_length
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Congelar encabezado + autofiltro
    def _freeze_and_filter(self, ws):
        if ws.max_row > 1 and ws.max_column > 0:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    # Contar severidades NVD
    def _count_severities(self, ports):
        severities = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

        for p in ports:
            for v in getattr(p, "vulnerabilities", []):
                sev = (v.severity or "").upper()
                if sev in severities:
                    severities[sev] += 1

        return severities

    # Pintar severidades por color
    def _apply_severity_colors(self, ws, severity_col: int, start_row: int = 2):
        """
        Aplica color de fondo a la columna de severidades.
        severity_col: índice de columna (1-based).
        """
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=severity_col)
            sev = str(cell.value or "").upper()
            color = self.SEVERITY_COLORS.get(sev)
            if color:
                cell.fill = PatternFill(start_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF")

    #   MODO SINGLE HOST
    def generate_host_excel(self, host, ports, risk_assessments, output_path: str) -> str:
        """
        Reporte detallado de un solo host:
            - Host Info
            - Ports
            - Vulnerabilities (NVD)
            - Risk Assessment
            - Pie chart de severidades
        """
        wb = Workbook()

        # HOJA 1: Información del Host
        ws = wb.active
        ws.title = "Host Info"

        ws.append(["Campo", "Valor"])
        ws.append(["IP Address", host.ip_address])
        ws.append(["Hostname", host.hostname or "N/A"])
        ws.append(["Sistema Operativo Detectado", host.os_detected or "N/A"])
        ws.append(["Fecha de Escaneo", host.scan_date.strftime("%Y-%m-%d %H:%M")])
        ws.append(["Total Puertos Escaneados", host.total_ports])
        ws.append(["Puertos Alto Riesgo", host.high_risk_count])

        self._style_header(ws, self.COLORS["header_blue"])
        self._auto_fit_columns(ws)
        self._freeze_and_filter(ws)

        # HOJA 2: Puertos Escaneados
        ws_ports = wb.create_sheet("Ports")
        ws_ports.append(["Puerto", "Protocolo", "Estado", "Servicio", "Versión", "Fecha"])

        for p in ports:
            ws_ports.append(
                [
                    p.port_number,
                    p.protocol.upper(),
                    p.status,
                    p.service_name or "N/A",
                    p.service_version or "N/A",
                    p.scanned_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )

        self._style_header(ws_ports, self.COLORS["header_gray"])
        self._auto_fit_columns(ws_ports)
        self._freeze_and_filter(ws_ports)

        # HOJA 3: Vulnerabilidades NVD (Landscape)
        ws_vuln = wb.create_sheet("Vulnerabilities")
        ws_vuln.append(["CVE", "CVSS", "Severidad", "Descripción", "Publicación"])

        ws_vuln.page_setup.orientation = "landscape"
        ws_vuln.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

        for p in ports:
            for v in p.vulnerabilities:
                ws_vuln.append(
                    [
                        v.cve_id,
                        v.cvss_score,
                        v.severity,
                        (v.description[:150] + "...") if v.description else "N/A",
                        v.published_date.strftime("%Y-%m-%d") if v.published_date else "N/A",
                    ]
                )

        self._style_header(ws_vuln, self.COLORS["header_red"])
        self._auto_fit_columns(ws_vuln)
        self._freeze_and_filter(ws_vuln)
        # Severidad está en la columna 3
        if ws_vuln.max_row > 1:
            self._apply_severity_colors(ws_vuln, severity_col=3)

        # HOJA 4: Evaluación de Riesgo
        ws_risk = wb.create_sheet("Risk Assessment")

        if risk_assessments:
            ra = risk_assessments[-1]
            ws_risk.append(["Campo", "Valor"])
            ws_risk.append(["Nivel de Riesgo", ra.risk_level])
            ws_risk.append(["Puntaje 0-100", ra.overall_risk_score])
            ws_risk.append(["Versión Modelo", ra.model_version])
            ws_risk.append(["Fecha Evaluada", ra.evaluated_at.strftime("%Y-%m-%d %H:%M")])
        else:
            ws_risk.append(["Sin evaluación de riesgo disponible"])

        self._style_header(ws_risk, self.COLORS["header_orange"])
        self._auto_fit_columns(ws_risk)
        self._freeze_and_filter(ws_risk)

        # HOJA 5: Gráfica PIE de Severidades
        ws_chart = wb.create_sheet("NVD Chart")
        ws_chart.append(["Severidad", "Cantidad"])

        severity_counts = self._count_severities(ports)
        for sev, count in severity_counts.items():
            ws_chart.append([sev, count])

        chart = PieChart()
        chart.title = "Distribución de Severidades NVD"

        # Columna 2: Cantidad (incluye título)
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=ws_chart.max_row)
        # Columna 1: Severidad (sin el encabezado)
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=ws_chart.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.width = 13
        chart.height = 9

        ws_chart.add_chart(chart, "D2")
        self._auto_fit_columns(ws_chart)

        wb.save(output_path)
        return output_path

    #   MODO MULTI HOST (reporte global)
    def generate_network_excel(self, hosts, output_path: str) -> str:
        """
        Genera un Excel con TODOS los hosts del usuario en una única hoja:
        - Una fila por (host, puerto, vulnerabilidad)
        - Ordenado por IP ascendente
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Report"

        ws.page_setup.orientation = "landscape"
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

        # Encabezados
        ws.append(
            [
                "Host IP",
                "Hostname",
                "OS",
                "Fecha Escaneo",
                "Puerto",
                "Protocolo",
                "Estado",
                "Servicio",
                "Versión",
                "CVE",
                "CVSS",
                "Severidad",
                "Descripción",
                "Publicación",
                "Nivel Riesgo",
                "Score Riesgo",
            ]
        )
        self._style_header(ws, self.COLORS["header_blue"])

        # Ordenar hosts por IP ascendente
        hosts_sorted = sorted(hosts, key=lambda h: str(h.ip_address))

        # Filas de datos
        for host in hosts_sorted:
            ports = getattr(host, "ports", []) or []
            risk_assessments = getattr(host, "risk_assessments", []) or []
            ra = risk_assessments[-1] if risk_assessments else None

            risk_level = ra.risk_level if ra else "N/A"
            risk_score = ra.overall_risk_score if ra else None

            if not ports:
                # host sin puertos registrados
                ws.append(
                    [
                        host.ip_address,
                        host.hostname or "N/A",
                        host.os_detected or "N/A",
                        host.scan_date.strftime("%Y-%m-%d %H:%M") if host.scan_date else "N/A",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        risk_level,
                        risk_score,
                    ]
                )
                continue

            for p in ports:
                vulns = getattr(p, "vulnerabilities", []) or []

                # Si no hay vulnerabilidades, una fila por puerto
                if not vulns:
                    ws.append(
                        [
                            host.ip_address,
                            host.hostname or "N/A",
                            host.os_detected or "N/A",
                            host.scan_date.strftime("%Y-%m-%d %H:%M") if host.scan_date else "N/A",
                            p.port_number,
                            p.protocol.upper(),
                            p.status,
                            p.service_name or "N/A",
                            p.service_version or "N/A",
                            None,
                            None,
                            None,
                            None,
                            None,
                            risk_level,
                            risk_score,
                        ]
                    )
                else:
                    # Una fila por vulnerabilidad
                    for v in vulns:
                        ws.append(
                            [
                                host.ip_address,
                                host.hostname or "N/A",
                                host.os_detected or "N/A",
                                host.scan_date.strftime("%Y-%m-%d %H:%M") if host.scan_date else "N/A",
                                p.port_number,
                                p.protocol.upper(),
                                p.status,
                                p.service_name or "N/A",
                                p.service_version or "N/A",
                                v.cve_id,
                                v.cvss_score,
                                v.severity,
                                (v.description[:150] + "...") if v.description else "N/A",
                                v.published_date.strftime("%Y-%m-%d") if v.published_date else "N/A",
                                risk_level,
                                risk_score,
                            ]
                        )

        # Ajustes finales
        self._auto_fit_columns(ws)
        self._freeze_and_filter(ws)
        # Severidad está en la columna 12
        if ws.max_row > 1:
            self._apply_severity_colors(ws, severity_col=12)

        wb.save(output_path)
        return output_path
